from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from .models import Assignment, Submission
from .forms import AssignmentForm, SubmissionForm, GradeForm
from courses.models import Course, Enrollment
from quiz.models import Quiz, QuizAttempt
from tests.models import Test, StudentResponse
from projects.models import ProjectSubmission
from attendance.models import AttendanceSession, AttendanceRecord
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
def assignment_create(request, course_pk):
    course = get_object_or_404(Course, pk=course_pk, teacher=request.user)
    if request.method == 'POST':
        form = AssignmentForm(request.POST, request.FILES, course=course)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.course = course
            assignment.save()
            form.save_m2m()
            messages.success(request, 'Assignment created successfully!')
            return redirect('course_detail', pk=course_pk)
    else:
        form = AssignmentForm(course=course)
    return render(request, 'assignments/assignment_form.html', {
        'form': form, 
        'course': course,
        'page_title': 'Create Assignment',
        'submit_label': 'Create',
    })


@login_required
def assignment_edit(request, pk):
    assignment = get_object_or_404(Assignment, pk=pk)
    course = assignment.course
    if course.teacher != request.user:
        messages.error(request, 'You do not have permission to edit this assignment.')
        return redirect('course_detail', pk=course.pk)
        
    if request.method == 'POST':
        form = AssignmentForm(request.POST, request.FILES, instance=assignment, course=course)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.save()
            form.save_m2m()
            messages.success(request, 'Assignment updated successfully!')
            return redirect('assignment_detail', pk=assignment.pk)
    else:
        form = AssignmentForm(instance=assignment, course=course)
    return render(request, 'assignments/assignment_form.html', {
        'form': form,
        'course': course,
        'page_title': 'Edit Assignment',
        'submit_label': 'Save Changes',
    })


@login_required
def assignment_detail(request, pk):
    assignment  = get_object_or_404(Assignment, pk=pk)
    
    # Access control: only assigned students can view
    if request.user.is_student():
        # Check if student is enrolled in course first
        is_enrolled = Enrollment.objects.filter(student=request.user, course=assignment.course).exists()
        if not is_enrolled:
            messages.error(request, 'You are not enrolled in this course.')
            return redirect('course_list')
        
        # Check assignment allocation
        if assignment.assigned_to.exists() and not assignment.assigned_to.filter(id=request.user.id).exists():
            messages.error(request, 'You are not assigned to this assignment.')
            return redirect('course_detail', pk=assignment.course.pk)

    submission  = None
    submissions = None

    if request.user.is_student():
        submission = Submission.objects.filter(
            assignment=assignment, student=request.user
        ).first()

    if request.user.is_teacher():
        submissions = assignment.submissions.all()

    # ── FIX 1: Robust deadline check using aware datetimes ──────
    now        = timezone.now()
    is_overdue = now > assignment.due_date

    return render(request, 'assignments/assignment_detail.html', {
        'assignment':  assignment,
        'submission':  submission,
        'submissions': submissions,
        'is_overdue':  is_overdue,
        'now':         now,
    })


@login_required
def assignment_submit(request, pk):
    assignment = get_object_or_404(Assignment, pk=pk)

    if not request.user.is_student():
        messages.error(request, 'Only students can submit assignments.')
        return redirect('assignment_detail', pk=pk)

    # Access control: only assigned students can submit
    if assignment.assigned_to.exists() and not assignment.assigned_to.filter(id=request.user.id).exists():
        messages.error(request, 'You are not assigned to this assignment.')
        return redirect('course_detail', pk=assignment.course.pk)

    # ── FIX 1: Hard deadline lock — server side check ───────────
    now = timezone.now()
    if now > assignment.due_date:
        messages.error(
            request,
            f'Submission closed! The deadline was '
            f'{assignment.due_date.strftime("%b %d, %Y at %H:%M")}. '
            f'Late submissions are not accepted.'
        )
        return redirect('assignment_detail', pk=pk)

    existing = Submission.objects.filter(
        assignment=assignment, student=request.user
    ).first()
    if existing:
        messages.warning(request, 'You have already submitted this assignment.')
        return redirect('assignment_detail', pk=pk)

    if request.method == 'POST':
        form = SubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            submission          = form.save(commit=False)
            submission.assignment = assignment
            submission.student    = request.user
            submission.save()
            messages.success(request, 'Assignment submitted successfully!')
            return redirect('assignment_detail', pk=pk)
    else:
        form = SubmissionForm()

    return render(request, 'assignments/submit_form.html', {
        'form': form, 'assignment': assignment,
    })


@login_required
def grade_submission(request, pk):
    submission = get_object_or_404(Submission, pk=pk)
    if not request.user.is_teacher():
        messages.error(request, 'Only teachers can grade submissions.')
        return redirect('dashboard')
    if request.method == 'POST':
        form = GradeForm(request.POST, instance=submission)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grade saved successfully!')
            return redirect('assignment_detail', pk=submission.assignment.pk)
    else:
        form = GradeForm(instance=submission)
    return render(request, 'assignments/grade_form.html', {
        'form': form, 'submission': submission,
    })


@login_required
def assignment_delete(request, pk):
    assignment = get_object_or_404(Assignment, pk=pk, course__teacher=request.user)
    course_pk  = assignment.course.pk
    if request.method == 'POST':
        assignment.delete()
        messages.success(request, 'Assignment deleted.')
        return redirect('course_detail', pk=course_pk)
    return render(request, 'assignments/assignment_confirm_delete.html', {
        'assignment': assignment,
    })


# ── FIX 3: Delete own submission (student) ──────────────────────
@login_required
def delete_submission(request, pk):
    submission = get_object_or_404(Submission, pk=pk, student=request.user)
    assignment_pk = submission.assignment.pk

    # Allow delete regardless of grade status as requested by UI needs

    if request.method == 'POST':
        submission.file.delete(save=False)
        submission.delete()
        messages.success(request, 'Submission deleted successfully.')
        return redirect('assignment_detail', pk=assignment_pk)

    return render(request, 'assignments/submission_confirm_delete.html', {
        'submission': submission,
    })


# ── Feature 7: Export CSV ────────────────────────────────────────
@login_required
def export_marks_csv(request, course_pk):
    course = get_object_or_404(Course, pk=course_pk, teacher=request.user)
    enrollments = Enrollment.objects.filter(course=course).select_related('student').order_by('student__username')
    
    assignments = list(Assignment.objects.filter(course=course).order_by('created_at'))
    quizzes = list(Quiz.objects.filter(course=course).order_by('created_at'))
    tests = list(Test.objects.filter(course=course).order_by('created_at'))

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{course.title}_full_marks.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Student Name', 'Username', 'Email',
        'Type', 'Title', 'Total Marks', 'Obtained Marks',
        'Percentage', 'Status'
    ])

    for enrollment in enrollments:
        student = enrollment.student
        
        # 1. Assignments
        for asm in assignments:
            sub = Submission.objects.filter(student=student, assignment=asm).first()
            score = sub.grade if sub else 0
            pct = f"{round((score / asm.total_marks) * 100, 1)}%" if asm.total_marks and score is not None else '0%'
            status = 'Graded' if sub and sub.grade is not None else ('Submitted' if sub else 'Not Submitted')
            writer.writerow([
                student.get_full_name() or student.username,
                student.username,
                student.email,
                asm.get_label_display(),
                asm.title,
                asm.total_marks,
                score if score is not None else 0,
                pct,
                status
            ])

        # 2. Quizzes
        for qz in quizzes:
            attempt = QuizAttempt.objects.filter(student=student, quiz=qz, is_complete=True).first()
            score = attempt.score if attempt else 0
            pct = f"{round((score / qz.total_marks) * 100, 1)}%" if qz.total_marks else '0%'
            status = 'Graded' if attempt else 'Not Attempted'
            writer.writerow([
                student.get_full_name() or student.username,
                student.username,
                student.email,
                'Quiz',
                qz.title,
                qz.total_marks,
                score,
                pct,
                status
            ])

        # 3. Tests
        for tst in tests:
            resp = StudentResponse.objects.filter(student=student, test=tst).first()
            score = resp.score if resp else 0
            pct = f"{round((score / tst.total_marks) * 100, 1)}%" if tst.total_marks else '0%'
            status = 'Graded' if resp else 'Missed'
            writer.writerow([
                student.get_full_name() or student.username,
                student.username,
                student.email,
                'Test',
                tst.title,
                tst.total_marks,
                score,
                pct,
                status
            ])

        # 4. Projects
        projects = ProjectSubmission.objects.filter(student=student, course=course)
        for prj in projects:
            score = prj.score
            pct = f"{round((score / prj.total_marks) * 100, 1)}%" if prj.total_marks and score is not None else ('0%' if score is not None else 'N/A')
            status = 'Graded' if score is not None else 'Submitted (Pending)'
            writer.writerow([
                student.get_full_name() or student.username,
                student.username,
                student.email,
                'Project',
                prj.title,
                prj.total_marks,
                score if score is not None else 'Pending',
                pct,
                status
            ])

        # 5. Attendance Summary
        total_sessions = AttendanceSession.objects.filter(course=course).count()
        if total_sessions > 0:
            present_count = AttendanceRecord.objects.filter(
                session__course=course, student=student, status='present'
            ).count()
            att_pct = f"{round((present_count / total_sessions) * 100, 1)}%"
            writer.writerow([
                student.get_full_name() or student.username,
                student.username,
                student.email,
                'Attendance',
                'Course Attendance',
                total_sessions,
                present_count,
                att_pct,
                'N/A'
            ])

    return response


# ── Feature 7: Export Excel ──────────────────────────────────────
@login_required
def export_marks_excel(request, course_pk):
    course = get_object_or_404(Course, pk=course_pk, teacher=request.user)
    enrollments = Enrollment.objects.filter(course=course).select_related('student').order_by('student__username')
    
    assignments = list(Assignment.objects.filter(course=course).order_by('created_at'))
    quizzes = list(Quiz.objects.filter(course=course).order_by('created_at'))
    tests = list(Test.objects.filter(course=course).order_by('created_at'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Full Marks Report'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'Student Name', 'Username', 'Email', 'Type', 'Title',
        'Total Marks', 'Obtained Marks', 'Percentage', 'Status'
    ]
    col_widths = [20, 15, 25, 12, 25, 12, 14, 12, 15]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    row_num = 2
    for enrollment in enrollments:
        student = enrollment.student
        
        # Collect all rows for this student
        student_rows = []
        
        # 1. Assignments
        for asm in assignments:
            sub = Submission.objects.filter(student=student, assignment=asm).first()
            score = sub.grade if sub else 0
            student_rows.append({
                'type': asm.get_label_display(),
                'title': asm.title,
                'total': asm.total_marks,
                'score': score if score is not None else 0,
                'status': 'Graded' if sub and sub.grade is not None else ('Submitted' if sub else 'Not Submitted')
            })

        # 2. Quizzes
        for qz in quizzes:
            attempt = QuizAttempt.objects.filter(student=student, quiz=qz, is_complete=True).first()
            score = attempt.score if attempt else 0
            student_rows.append({
                'type': 'Quiz',
                'title': qz.title,
                'total': qz.total_marks,
                'score': score,
                'status': 'Graded' if attempt else 'Not Attempted'
            })

        # 3. Tests
        for tst in tests:
            resp = StudentResponse.objects.filter(student=student, test=tst).first()
            score = resp.score if resp else 0
            student_rows.append({
                'type': 'Test',
                'title': tst.title,
                'total': tst.total_marks,
                'score': score,
                'status': 'Graded' if resp else 'Missed'
            })

        # 4. Projects
        projects = ProjectSubmission.objects.filter(student=student, course=course)
        for prj in projects:
            score = prj.score
            student_rows.append({
                'type': 'Project',
                'title': prj.title,
                'total': prj.total_marks,
                'score': score if score is not None else 'Pending',
                'status': 'Graded' if score is not None else 'Submitted (Pending)'
            })

        # 5. Attendance Summary
        total_sessions = AttendanceSession.objects.filter(course=course).count()
        if total_sessions > 0:
            present_count = AttendanceRecord.objects.filter(
                session__course=course, student=student, status='present'
            ).count()
            att_pct = f"{round((present_count / total_sessions) * 100, 1)}%"
            student_rows.append({
                'type': 'Attendance',
                'title': 'Course Attendance',
                'total': total_sessions,
                'score': present_count,
                'pct': att_pct,
                'status': 'N/A'
            })

        # Write rows to sheet
        for r_data in student_rows:
            total = r_data['total']
            score = r_data['score']
            status = r_data['status']
            
            # Use calculated pct if provided (for attendance), else calculate
            pct = r_data.get('pct')
            if not pct:
                if isinstance(score, (int, float)) and total:
                    pct = f"{round((score / total) * 100, 1)}%"
                elif status == 'Graded':
                    pct = '0%'
                else:
                    pct = 'N/A'

            row_values = [
                student.get_full_name() or student.username,
                student.username,
                student.email,
                r_data['type'],
                r_data['title'],
                total,
                score,
                pct,
                status
            ]
            
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                # Apply conditional color based on status/type
                if r_data['type'] == 'Attendance':
                    cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid') # Light Blue
                elif 'Graded' in status:
                    cell.fill = green_fill
                elif 'Submitted' in status or 'Pending' in status:
                    cell.fill = amber_fill
                else:
                    cell.fill = red_fill
                
                cell.alignment = Alignment(vertical='center')
            
            row_num += 1

    # ── Weekly Analysis Sheet ─────────────────────────────────────────
    ws_weekly = wb.create_sheet(title='Weekly Analysis')
    
    headers_weekly = [
        'Student Name', 'Username', 'Email', 'Week', 'Date Range',
        'Assignments Submitted', 'Quizzes Attempted', 'Tests Taken', 'Attendance Rate'
    ]
    col_widths_weekly = [20, 15, 25, 12, 25, 22, 18, 15, 18]
    
    for col_num, (header, width) in enumerate(zip(headers_weekly, col_widths_weekly), 1):
        cell = ws_weekly.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws_weekly.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
    ws_weekly.row_dimensions[1].height = 25
    
    # Calculate weeks starting from course.start_date or course.created_at
    course_start = course.start_date or course.created_at
    now = timezone.now()
    
    from datetime import timedelta
    weeks = []
    current_start = course_start
    week_num = 1
    
    while current_start < now:
        current_end = current_start + timedelta(days=7)
        weeks.append({
            'num': week_num,
            'start': current_start,
            'end': current_end,
        })
        current_start = current_end
        week_num += 1
        if week_num > 24: # Limit to prevent infinite loop
            break
            
    # Fallback to last 4 weeks if no weeks generated
    if not weeks:
        today = timezone.now()
        for i in range(4):
            end_date = today - timedelta(days=7 * i)
            start_date = end_date - timedelta(days=7)
            weeks.append({
                'num': 4 - i,
                'start': start_date,
                'end': end_date,
            })
        weeks.sort(key=lambda x: x['num'])
        
    row_num_weekly = 2
    for enrollment in enrollments:
        student = enrollment.student
        student_name = student.get_full_name() or student.username
        
        for wk in weeks:
            start_date = wk['start']
            end_date = wk['end']
            
            # 1. Assignments
            asm_submitted = Submission.objects.filter(
                student=student,
                assignment__course=course,
                submitted_at__range=(start_date, end_date)
            ).count()
            
            # 2. Quizzes
            quiz_completed = QuizAttempt.objects.filter(
                student=student,
                quiz__course=course,
                is_complete=True,
                finished_at__range=(start_date, end_date)
            ).count()
            
            # 3. Tests
            test_completed = StudentResponse.objects.filter(
                student=student,
                test__course=course,
                submitted_at__range=(start_date, end_date)
            ).count()
            
            # 4. Attendance
            total_sessions = AttendanceSession.objects.filter(
                course=course,
                date__range=(start_date.date(), end_date.date())
            ).count()
            if total_sessions > 0:
                present_count = AttendanceRecord.objects.filter(
                    student=student,
                    session__course=course,
                    status='present',
                    session__date__range=(start_date.date(), end_date.date())
                ).count()
                att_pct = f"{round((present_count / total_sessions) * 100, 1)}%"
            else:
                att_pct = 'N/A'
                
            row_values = [
                student_name,
                student.username,
                student.email,
                f"Week {wk['num']}",
                f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
                asm_submitted,
                quiz_completed,
                test_completed,
                att_pct
            ]
            
            for col_num, value in enumerate(row_values, 1):
                cell = ws_weekly.cell(row=row_num_weekly, column=col_num, value=value)
                cell.alignment = Alignment(vertical='center')
                
            row_num_weekly += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{course.title}_full_marks.xlsx"'
    wb.save(response)
    return response
